import requests
from msal import ConfidentialClientApplication
import pandas as pd
import json
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys

# === CONFIG ===
tenant_id = "114c8106-747f-4cc7-870e-8712e6c23b18"
client_id = "b357e50c-c5ef-484d-84df-fe470fe76528"
client_secret = "JAZ8Q~xlY-EDlgbLtgJaqjPNAjsHfYFavwxbkdjE"

site_hostname = "juggernautenterprises.sharepoint.com"
site_path = "/sites/TestCloud"
doc_library = "TestClient"
fy_year = "TestClient_FY25"


def get_token():
    app = ConfidentialClientApplication(client_id, authority=f"https://login.microsoftonline.com/{tenant_id}", client_credential=client_secret)
    token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in token_response:
        raise Exception("Failed to acquire access token")
    return token_response["access_token"]


def get_drive_id(headers):
    site_url = f"https://graph.microsoft.com/v1.0/sites/{site_hostname}:{site_path}"
    site_resp = requests.get(site_url, headers=headers)
    site_resp.raise_for_status()
    site_id = site_resp.json()["id"]
    drives_resp = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives", headers=headers)
    drives_resp.raise_for_status()
    for d in drives_resp.json()["value"]:
        if d["name"] == doc_library:
            return d["id"]
    raise Exception(f"Library '{doc_library}' not found")


def download_file(headers, drive_id, file_name, folder_name):
    download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{fy_year}/{folder_name}/{file_name}:/content"
    resp = requests.get(download_url, headers=headers)
    resp.raise_for_status()
    return resp.content


def upload_file(headers, drive_id, file_bytes, remote_name, folder_name):
    upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{fy_year}/{folder_name}/{remote_name}:/content"
    resp = requests.put(upload_url, headers=headers, data=file_bytes.getvalue())
    resp.raise_for_status()
    return resp.json().get("webUrl")


def process_reconciliation(json_bytes, excel_bytes, amt_col_name):
    # Normalize incoming column name
    amt_col_name = str(amt_col_name or "").strip()
    data = json.loads(json_bytes.decode("utf-8"))
    if isinstance(data, dict):
        data = [data]
    closing_sum = round(sum(item.get("closing_balance", 0.0) for item in data if item.get("fs_sub_line_id") == 20015), 2)
    df = pd.read_excel(BytesIO(excel_bytes))
    df.columns = df.columns.str.strip()
    if amt_col_name not in df.columns:
        raise ValueError(f"Column '{amt_col_name}' not found in Excel file.")
    excel_sum = round(df[amt_col_name].sum(), 2)
    difference = excel_sum - closing_sum
    output = {"Reconciliation": [
        {"Particulars": "As per CWIP Register (A)", "Amount": excel_sum},
        {"Particulars": "As per Sublead (B)", "Amount": closing_sum},
        {"Particulars": "Difference (C)", "Amount": difference},
        {"Particulars": "User Adjustments", "Details": []},
        {"Particulars": "Net Difference", "Amount": difference}
    ]}
    return BytesIO(json.dumps(output, indent=4).encode("utf-8"))


def process_ageing_excel(excel_bytes, date_col_name, cutoff_date):
    # Normalize incoming column name
    date_col_name = str(date_col_name or "").strip()
    wb = openpyxl.load_workbook(BytesIO(excel_bytes))
    ws = wb.worksheets[0]
    max_col = ws.max_column
    max_row = ws.max_row
    date_col_letter = None
    for col in range(1, max_col + 1):
        if str(ws.cell(row=1, column=col).value).strip() == date_col_name:
            date_col_letter = get_column_letter(col)
            break
    if not date_col_letter:
        raise ValueError(f"Column '{date_col_name}' not found in first sheet")
    table_start_col = max_col + 6
    guide_data = [["Ageing Guide", "Ageing Bracket"], [-1000, "Less than 1 year"], [366, "1-2 Years"], [731, "2-3 Years"], [1096, "More than 3 years"]]
    for r, row_data in enumerate(guide_data, start=1):
        for c, val in enumerate(row_data, start=table_start_col):
            ws.cell(row=r, column=c, value=val)
    new_col1 = max_col + 1
    new_col2 = max_col + 2
    cutoff_date_dt = datetime.strptime(cutoff_date, "%m/%d/%Y").date()
    ws.cell(row=1, column=new_col1, value=cutoff_date_dt)
    ws.cell(row=1, column=new_col2, value="ET Ageing")
    for row in range(2, max_row + 1):
        ws.cell(row=row, column=new_col1).value = f"=${get_column_letter(new_col1)}$1-{date_col_letter}{row}+1"
        ws.cell(row=row, column=new_col2).value = (f"=VLOOKUP({get_column_letter(new_col1)}{row}," f"${get_column_letter(table_start_col)}$2:${get_column_letter(table_start_col+1)}$5,2,1)")
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def jugg(excel_file, cutoff_date, columns):
    amt_col_name, date_col_name = columns
    # Normalize column names early
    amt_col_name = str(amt_col_name or "").strip()
    date_col_name = str(date_col_name or "").strip()
    access_token = get_token()
    headers = {"Authorization": f"Bearer {access_token}"}
    drive_id = get_drive_id(headers)
    json_content = download_file(headers, drive_id, "FinData_LedgerMapping.json", "juggernaut")
    excel_content = download_file(headers, drive_id, excel_file, "client")
    json_buffer = process_reconciliation(json_content, excel_content, amt_col_name)
    excel_buffer = process_ageing_excel(excel_content, date_col_name, cutoff_date)
    json_url = upload_file(headers, drive_id, json_buffer, "Execution_PPE_CompletenessCheck.json", "juggernaut")
    excel_url = upload_file(headers, drive_id, excel_buffer, "Execution_PPE_CwipAnalysis.xlsx", "client")
    db_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{fy_year}/juggernaut/db.json:/content"
    db_resp = requests.get(db_url, headers=headers)
    db_data = db_resp.json() if db_resp.status_code == 200 else {"juggernaut": [], "client": [], "tools": [], "rbin": []}
    db_data["client"].append({"name": "Execution_PPE_CwipAnalysis.xlsx", "url": excel_url, "reference": "Execution PPE CWIP Ageing"})
    db_data["juggernaut"].append({"name": "Execution_PPE_CompletenessCheck.json", "url": json_url, "reference": ""})
    requests.put(db_url, headers=headers, data=json.dumps(db_data, indent=4).encode("utf-8"))
    print("✅ Uploaded Excel:", excel_url)
    print("✅ Uploaded JSON:", json_url)
    print("✅ db.json updated")
    return {"success": True, "excel_url": excel_url, "json_url": json_url}


def main():
    payload = None
    if len(sys.argv) >= 3:
        cfg_path = sys.argv[2]
        try:
            with open(cfg_path, 'r', encoding='utf-8') as f:
                payload = json.load(f)
        except Exception as e:
            print(json.dumps({"success": False, "error": f"Failed to read config: {e}"}))
            sys.exit(1)
    if not payload:
        print(json.dumps({"success": False, "error": "Missing config payload"}))
        sys.exit(1)
    opts = payload.get("options", payload)
    excel_file = opts.get("excel_file", "")
    cutoff_date = opts.get("cutoff_date", "")
    columns = opts.get("columns", [])
    try:
        res = jugg(excel_file, cutoff_date, columns)
        print(json.dumps(res))
        sys.exit(0)
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
