import requests
from msal import ConfidentialClientApplication
import pandas as pd
import json
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === CONFIG ===
tenant_id = "114c8106-747f-4cc7-870e-8712e6c23b18"
client_id = "b357e50c-c5ef-484d-84df-fe470fe76528"
client_secret = "JAZ8Q~xlY-EDlgbLtgJaqjPNAjsHfYFavwxbkdjE"

site_hostname = "juggernautenterprises.sharepoint.com"
site_path = "/sites/TestCloud"
doc_library = "TestClient"
fy_year = "TestClient_FY25"


# === AUTH ===
def get_token():
    app = ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        client_credential=client_secret
    )
    token_response = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in token_response:
        raise Exception("Failed to acquire access token")
    return token_response["access_token"]


def get_drive_id(headers):
    site_url = f"https://graph.microsoft.com/v1.0/sites/{site_hostname}:{site_path}"
    site_resp = requests.get(site_url, headers=headers)
    site_resp.raise_for_status()
    site_id = site_resp.json()["id"]

    drives_resp = requests.get(f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives", headers=headers)
    drives_resp.raise_for_status()
    drives = drives_resp.json()["value"]

    for d in drives:
        if d["name"] == doc_library:
            return d["id"]
    raise Exception(f"Library '{doc_library}' not found")


def download_file(headers, drive_id, file_name, folder_name):
    download_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{fy_year}/{folder_name}/{file_name}:/content"
    resp = requests.get(download_url, headers=headers)
    resp.raise_for_status()
    return resp.content


def upload_file(headers, drive_id, local_path, remote_name, folder_name):
    upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{fy_year}/{folder_name}/{remote_name}:/content"
    with open(local_path, "rb") as f:
        data = f.read()
    resp = requests.put(upload_url, headers=headers, data=data)
    resp.raise_for_status()
    return resp.json().get("webUrl")


# === MoM INCREMENT ANALYSIS ===
def run_mom_increment(pay_registrar_content, column_map_content, display_cols, calc_cols, increment_month, output_path):
    cm_data = json.loads(column_map_content)
    employee_code = cm_data["column_map"]["employee_code"]
    pay_month = cm_data["column_map"]["pay_month"]

    # Load data
    df = pd.read_excel(BytesIO(pay_registrar_content))

    # Strip time from datetime columns
    for col in df.select_dtypes(include=["datetime64[ns]"]).columns:
        df[col] = df[col].dt.date

    # Ensure month column is datetime
    df[pay_month] = pd.to_datetime(df[pay_month], errors="coerce")

    # Calculate total per row
    df["Total_Calc"] = df[calc_cols].sum(axis=1)

    # Aggregate totals by employee & month
    totals = df.groupby([employee_code, pay_month], as_index=False)["Total_Calc"].sum()

    # Employee info
    agg_dict = {col: "first" for col in display_cols if col != employee_code}
    emp_info = df.groupby(employee_code, as_index=False).agg(agg_dict)

    # Pivot totals to months
    pivot = totals.pivot(index=employee_code, columns=pay_month, values="Total_Calc")
    pivot.columns = [pd.to_datetime(c).strftime("%b-%y") for c in pivot.columns]

    pivot_df = emp_info.merge(pivot.reset_index(), on=employee_code, how="left")

    # Identify pre/post increment months
    month_list = [col for col in pivot_df.columns if col not in display_cols]
    if increment_month not in month_list:
        raise ValueError(f"Increment month '{increment_month}' not found in data. Available: {month_list}")

    inc_index = month_list.index(increment_month)
    pre_months = month_list[:inc_index]
    post_months = month_list[inc_index:]

    # Build dataframe
    separator_col = [''] * len(pivot_df)
    combined_df = pd.concat(
        [
            pivot_df[display_cols + pre_months],
            pd.DataFrame(columns=["Pre Average", "Pre StdDev", "Pre Variance %"]),
            pd.DataFrame({'': separator_col}),
            pivot_df[post_months],
            pd.DataFrame(columns=["Post Average", "Post StdDev", "Post Variance %"])
        ],
        axis=1
    )

    # Save without formulas first
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        combined_df.to_excel(writer, sheet_name="MoM Analysis", index=False)

    # Insert formulas
    wb = load_workbook(output_path)
    ws = wb["MoM Analysis"]

    display_n = len(display_cols)
    pre_n = len(pre_months)
    pre_avg_col = display_n + pre_n + 1
    pre_std_col = pre_avg_col + 1
    pre_var_col = pre_std_col + 1
    sep_col = pre_var_col + 1
    post_start_col = sep_col + 1
    post_n = len(post_months)
    post_avg_col = post_start_col + post_n
    post_std_col = post_avg_col + 1
    post_var_col = post_std_col + 1

    ws.cell(row=1, column=pre_avg_col, value="Pre Average")
    ws.cell(row=1, column=pre_std_col, value="Pre StdDev")
    ws.cell(row=1, column=pre_var_col, value="Pre Variance %")
    ws.cell(row=1, column=post_avg_col, value="Post Average")
    ws.cell(row=1, column=post_std_col, value="Post StdDev")
    ws.cell(row=1, column=post_var_col, value="Post Variance %")

    def std_formula(range_str):
        return f"IFERROR(STDEV.P({range_str}),STDEVP({range_str}))"

    for r in range(2, ws.max_row + 1):
        if pre_n > 0:
            pre_range = f"{get_column_letter(display_n + 1)}{r}:{get_column_letter(display_n + pre_n)}{r}"
            ws.cell(row=r, column=pre_avg_col, value=f"=ROUND(AVERAGE({pre_range}),2)")
            ws.cell(row=r, column=pre_std_col, value=f"=ROUND({std_formula(pre_range)},2)")
            ws.cell(row=r, column=pre_var_col, value=f"=ROUND(IFERROR({get_column_letter(pre_std_col)}{r}/{get_column_letter(pre_avg_col)}{r},0),2)")

        if post_n > 0:
            post_range = f"{get_column_letter(post_start_col)}{r}:{get_column_letter(post_start_col + post_n - 1)}{r}"
            ws.cell(row=r, column=post_avg_col, value=f"=ROUND(AVERAGE({post_range}),2)")
            ws.cell(row=r, column=post_std_col, value=f"=ROUND({std_formula(post_range)},2)")
            ws.cell(row=r, column=post_var_col, value=f"=ROUND(IFERROR({get_column_letter(post_std_col)}{r}/{get_column_letter(pre_avg_col)}{r},0),2)")

    wb.save(output_path)


# === MAIN ===
def jugg(pay_registrar_name, display_cols, calc_cols, increment_month):
    access_token = get_token()
    headers = {"Authorization": f"Bearer {access_token}"}
    drive_id = get_drive_id(headers)

    # Download files
    pay_registrar_content = download_file(headers, drive_id, pay_registrar_name, "client")
    column_map_content = download_file(headers, drive_id, "Execution_Payroll_PRColumnMap.json", "juggernaut")

    output_path = "Execution_Payroll_MomIncrement.xlsx"
    run_mom_increment(pay_registrar_content, column_map_content, display_cols, calc_cols, increment_month, output_path)

    # Upload file
    remote_name = "Execution_Payroll_MomIncrement.xlsx"
    file_web_url = upload_file(headers, drive_id, output_path, remote_name, "client")
    print("✅ Uploaded file:", file_web_url)

    # Update db.json
    db_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{fy_year}/juggernaut/db.json:/content"
    db_resp = requests.get(db_url, headers=headers)
    if db_resp.status_code == 200:
        db_data = db_resp.json()
    else:
        db_data = {"juggernaut": [], "client": [], "tools": [], "rbin": []}

    new_entry = {"name": remote_name, "url": file_web_url, "reference": "Execution Payroll MoM Increment"}
    db_data.setdefault("client", []).append(new_entry)

    updated_db_content = json.dumps(db_data, indent=4)
    db_upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{fy_year}/juggernaut/db.json:/content"
    db_upload_resp = requests.put(db_upload_url, headers=headers, data=updated_db_content.encode("utf-8"))
    db_upload_resp.raise_for_status()

    # Emit a single-line JSON for the frontend listener
    print(json.dumps({"success": True, "file_web_url": file_web_url}))


if __name__ == "__main__":
    # Example manual run (defaults)
    pay_registrar = "Pay Registrar.xlsx"
    display_cols = ['Pernr', 'Employee Name', 'DOJ', 'DOL']
    calc_cols = ['BASIC', 'H R A']
    increment_month = 'Nov-24'
    jugg(pay_registrar, display_cols, calc_cols, increment_month)


