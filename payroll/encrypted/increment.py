import pandas as pd
import json
import os
from openpyxl import load_workbook

# --- INPUTS ---
cy_file = r"C:\Users\shez8\Desktop\Juggernaut\Execution\1 Payroll\Inputs\CTC Report.xlsx"
cy_sheet = "Data"
py_file = r"C:\Users\shez8\Desktop\Juggernaut\Execution\1 Payroll\Inputs\CTC Report Prev year.xlsx"
py_sheet = "Data"
output_file = r"C:\Users\shez8\Desktop\Juggernaut\Execution\1 Payroll\Outputs\Mom Increment.xlsx"
incr_columns = ["Emp. No", "Emp. Name", "DOJ", "Department"]
cols_to_sum = ["Monthly CTC"]
reconcile_input = 0
output_path = r"C:\Users\shez8\Desktop\Juggernaut\Execution\1 Payroll\Outputs\Increment_Analysis.json"

def jugg(
    cy_file,
    cy_sheet,
    py_file,
    py_sheet,
    output_file,
    incr_columns,   # list of column names
    cols_to_sum,
    reconcile_input,
    output_path
):
    empno_col, name_col, doj_col, designation_col = incr_columns
    # --- LOAD FILES ---
    cy_df = pd.read_excel(cy_file, sheet_name=cy_sheet)
    py_df = pd.read_excel(py_file, sheet_name=py_sheet)

    # Normalize column names
    cy_df.columns = cy_df.columns.str.strip()
    py_df.columns = py_df.columns.str.strip()

    # Merge on Employee No
    merged_df = pd.merge(
        cy_df,
        py_df,
        left_on=empno_col,
        right_on=empno_col,
        suffixes=("_CY", "_PY")
    )

    # --- CALCULATIONS ---
    merged_df["As per CY"] = merged_df[[col + "_CY" for col in cols_to_sum]].sum(axis=1)
    merged_df["As per PY"] = merged_df[[col + "_PY" for col in cols_to_sum]].sum(axis=1)
    merged_df["Increment"] = merged_df["As per CY"] - merged_df["As per PY"]

    merged_df["Increment % Num"] = merged_df.apply(
        lambda x: (x["Increment"] / x["As per PY"]) if x["As per PY"] != 0 else 0,
        axis=1
    )

    average_increment = merged_df['Increment % Num'].mean()

    merged_df["Increment %"] = merged_df["Increment % Num"].apply(
        lambda v: round(v, 2)
    )

    # --- FINAL OUTPUT DATA (unchanged logic) ---
    final_df = merged_df[
        [empno_col, name_col + "_CY", doj_col + "_CY", designation_col + "_CY",
        "As per CY", "As per PY", "Increment", "Increment %"]
    ].copy()

    final_df.rename(columns={
        empno_col: "Employee Code",
        name_col + "_CY": "Employee Name",
        doj_col + "_CY": "DOJ",
        designation_col + "_CY": "Designation"
    }, inplace=True)

    final_df["DOJ"] = pd.to_datetime(final_df["DOJ"]).dt.strftime("%d-%m-%Y")

    # --- SAVE: write only the table with pandas, then write the Average on top with openpyxl ---
    # Use if_sheet_exists="replace" so reruns replace the sheet
    if os.path.exists(output_file):
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            final_df.to_excel(
                writer,
                sheet_name="Increment Analysis",
                index=False,
                startrow=3  # keep table starting at row 4 (Excel) as before
            )
    else:
        # create the workbook and sheet if it doesn't exist
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
            final_df.to_excel(
                writer,
                sheet_name="Increment Analysis",
                index=False,
                startrow=3
            )

    # Now open the sheet and write Average at the top (A1 header, B1 value)
    wb = load_workbook(output_file)
    ws = wb["Increment Analysis"]

    ws["A1"] = "Average Increment %"
    # Keep the same rounding you used earlier
    ws["B1"] = round(average_increment, 2)

    wb.save(output_file)

    # --- SUMMARY JSON ---
    total_before_delete = len(cy_df)
    total_after_merge = len(merged_df)

    particulars = {
        "As per Increment analysis": total_after_merge,
        "As per CTC Report": total_before_delete,
        "User reconciliations": reconcile_input,
        "Difference": total_after_merge - total_before_delete,
        "Average Increment %": round(average_increment * 100, 2),
        "Net Difference": (total_after_merge - total_before_delete) - reconcile_input
    }

    with open(output_path, "w") as f:
        json.dump(particulars, f, indent=4)
