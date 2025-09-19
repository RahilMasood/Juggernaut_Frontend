import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- INPUTS ---
file_path = r"C:\Users\shez8\Desktop\Juggernaut\Execution\1 Payroll\Inputs\Pay Registrar.xlsx"
sheet_name = "Consol AIC"
display_cols = ['Pernr', 'Employee Name', 'DOJ', 'DOL']
calc_cols = ['BASIC', 'H R A']
increment_month = 'Nov-24'
column_map = r"C:\Users\shez8\Desktop\Juggernaut\Execution\1 Payroll\Outputs\column_map.json"
output_path = r"C:\Users\shez8\Desktop\Juggernaut\Execution\1 Payroll\Outputs\Mom Increment.xlsx"

def jugg(file_path, sheet_name, column_map, output_path, display_cols, calc_cols, increment_month):
    with open(column_map, "r") as f:
        cm_data = json.load(f)
    employee_code = cm_data["column_map"]["employee_code"]
    pay_month = cm_data["column_map"]["pay_month"]

    # Load data
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Strip time from datetime columns
    for col in df.select_dtypes(include=["datetime64[ns]"]).columns:
        df[col] = df[col].dt.date

    # Ensure month column is datetime
    df[pay_month] = pd.to_datetime(df[pay_month], errors="coerce")

    # Calculate total per row
    df["Total_Calc"] = df[calc_cols].sum(axis=1)

    # Aggregate totals by employee & month
    totals = df.groupby([employee_code, pay_month], as_index=False)["Total_Calc"].sum()

    # Get employee info dynamically
    agg_dict = {col: "first" for col in display_cols if col != employee_code}
    emp_info = df.groupby(employee_code, as_index=False).agg(agg_dict)

    # Pivot totals to months
    pivot = totals.pivot(index=employee_code, columns=pay_month, values="Total_Calc")

    # Format month columns
    pivot.columns = [pd.to_datetime(c).strftime("%b-%y") for c in pivot.columns]

    # Merge employee info back
    pivot_df = emp_info.merge(pivot.reset_index(), on=employee_code, how="left")

    # Identify pre/post increment months
    month_list = [col for col in pivot_df.columns if col not in display_cols]
    if increment_month not in month_list:
        raise ValueError(f"Increment month '{increment_month}' not found in data. Available: {month_list}")

    inc_index = month_list.index(increment_month)
    pre_months = month_list[:inc_index]
    post_months = month_list[inc_index:]

    # Build dataframe in desired order
    separator_col = [''] * len(pivot_df)
    combined_df = pd.concat(
        [
            pivot_df[display_cols + pre_months],   # Display + pre months
            pd.DataFrame(columns=["Pre Average", "Pre StdDev", "Pre Variance %"]),  # Pre stats placeholder
            pd.DataFrame({'': separator_col}),    # Separator
            pivot_df[post_months],                 # Post months
            pd.DataFrame(columns=["Post Average", "Post StdDev", "Post Variance %"])  # Post stats placeholder
        ],
        axis=1
    )

    # Save without formulas first
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        combined_df.to_excel(writer, sheet_name="MoM Analysis", index=False)

    # Re-open with openpyxl to insert formulas
    wb = load_workbook(output_path)
    ws = wb["MoM Analysis"]

    # Locate column indexes
    display_n = len(display_cols)
    pre_n = len(pre_months)
    pre_avg_col = display_n + pre_n + 1
    pre_std_col = pre_avg_col + 1
    pre_var_col = pre_std_col + 1

    sep_col = pre_var_col + 1
    post_start_col = sep_col + 1
    post_n = len(post_months)
    post_avg_col = post_start_col + post_n
    post_std_col = post_avg_col + 1
    post_var_col = post_std_col + 1

    # Write headers (already written by pandas, but ensure they're correct)
    ws.cell(row=1, column=pre_avg_col, value="Pre Average")
    ws.cell(row=1, column=pre_std_col, value="Pre StdDev")
    ws.cell(row=1, column=pre_var_col, value="Pre Variance %")
    ws.cell(row=1, column=post_avg_col, value="Post Average")
    ws.cell(row=1, column=post_std_col, value="Post StdDev")
    ws.cell(row=1, column=post_var_col, value="Post Variance %")

    # Helper for STDDEV formula
    def std_formula(range_str):
        return f"IFERROR(STDEV.P({range_str}),STDEVP({range_str}))"

    # Insert formulas for each row
    for r in range(2, ws.max_row + 1):
        if pre_n > 0:
            pre_range = f"{get_column_letter(display_n + 1)}{r}:{get_column_letter(display_n + pre_n)}{r}"
            ws.cell(row=r, column=pre_avg_col, value=f"=ROUND(AVERAGE({pre_range}),2)")
            ws.cell(row=r, column=pre_std_col, value=f"=ROUND({std_formula(pre_range)},2)")
            ws.cell(row=r, column=pre_var_col, value=f"=ROUND(IFERROR({get_column_letter(pre_std_col)}{r}/{get_column_letter(pre_avg_col)}{r},0),2)")

        if post_n > 0:
            post_range = f"{get_column_letter(post_start_col)}{r}:{get_column_letter(post_start_col + post_n - 1)}{r}"
            ws.cell(row=r, column=post_avg_col, value=f"=ROUND(AVERAGE({post_range}),2)")
            ws.cell(row=r, column=post_std_col, value=f"=ROUND({std_formula(post_range)},2)")
            ws.cell(row=r, column=post_var_col, value=f"=ROUND(IFERROR({get_column_letter(post_std_col)}{r}/{get_column_letter(pre_avg_col)}{r},0),2)")

    # Save final workbook
    wb.save(output_path)
    print(f"✅ Output with formulas saved to {output_path}")
